"""Excel export service for segmentation analysis data"""

import io
from datetime import datetime
from typing import Optional

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.inference import get_prediction_statistics, get_lr_statistics


def create_excel_report(
    predictions: np.ndarray,
    metadata: dict,
    voxel_dims: Optional[tuple] = None,
    scale_factors: Optional[tuple] = None,
) -> io.BytesIO:
    """
    Create an Excel report with segmentation analysis data.

    Args:
        predictions: Prediction mask (H, W) or (H, W, D)
        metadata: File metadata
        voxel_dims: Voxel dimensions in mm
        scale_factors: Scale factors for display upscaling correction

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )

    def style_header(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Filename", metadata.get("filename", "Unknown")),
        ("Dimensions", " x ".join(str(d) for d in metadata.get("dimensions", []))),
        ("Number of Slices", metadata.get("num_slices", 1)),
    ]
    if voxel_dims:
        summary_data.append(("Voxel Size (mm)", f"{voxel_dims[0]:.3f} x {voxel_dims[1]:.3f} x {voxel_dims[2]:.3f}"))

    ws_summary.append(["Property", "Value"])
    style_header(ws_summary, 1, 2)
    for prop, val in summary_data:
        ws_summary.append([prop, val])

    for row in ws_summary.iter_rows(min_row=2, max_col=2):
        for cell in row:
            cell.border = thin_border
    auto_width(ws_summary)

    # --- Sheet 2: Overall Statistics ---
    ws_stats = wb.create_sheet("Overall Statistics")
    overall_stats = get_prediction_statistics(predictions, voxel_dims, scale_factors)

    has_volume = voxel_dims is not None
    headers = ["Class ID", "Class Name", "Pixel Count", "Percentage (%)"]
    if has_volume:
        headers.extend(["Volume (mm3)", "Volume (cm3)"])

    ws_stats.append(headers)
    style_header(ws_stats, 1, len(headers))

    for s in overall_stats:
        if s["class_id"] == 0:
            continue
        row = [s["class_id"], s["class_name"], s["pixel_count"], s["percentage"]]
        if has_volume:
            row.extend([s.get("volume_mm3", ""), s.get("volume_cm3", "")])
        ws_stats.append(row)

    for row in ws_stats.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    auto_width(ws_stats)

    # --- Sheet 3: Left-Right Comparison ---
    ws_lr = wb.create_sheet("Left-Right Comparison")
    lr_stats = get_lr_statistics(predictions, voxel_dims, scale_factors)

    lr_headers = ["Class Name", "Left %", "Right %"]
    if has_volume:
        lr_headers.extend(["Left Volume (mm3)", "Right Volume (mm3)"])

    ws_lr.append(lr_headers)
    style_header(ws_lr, 1, len(lr_headers))

    # Build lookup maps
    left_map = {s["class_id"]: s for s in lr_stats["left"]}
    right_map = {s["class_id"]: s for s in lr_stats["right"]}
    all_ids = set(left_map.keys()) | set(right_map.keys())

    for cid in sorted(all_ids):
        if cid == 0:
            continue
        left = left_map.get(cid, {})
        right = right_map.get(cid, {})
        name = left.get("class_name") or right.get("class_name", f"Class {cid}")
        row = [
            name,
            left.get("percentage", 0),
            right.get("percentage", 0),
        ]
        if has_volume:
            row.extend([
                left.get("volume_mm3", 0),
                right.get("volume_mm3", 0),
            ])
        ws_lr.append(row)

    for row in ws_lr.iter_rows(min_row=2, max_col=len(lr_headers)):
        for cell in row:
            cell.border = thin_border
    auto_width(ws_lr)

    # --- Sheet 4: Per-Slice Breakdown (NIfTI only) ---
    is_3d = len(predictions.shape) == 3
    if is_3d:
        ws_slices = wb.create_sheet("Per-Slice Breakdown")
        slice_headers = ["Slice", "Class Name", "Pixel Count", "Percentage (%)"]
        if has_volume:
            slice_headers.extend(["Volume (mm3)", "Volume (cm3)"])

        ws_slices.append(slice_headers)
        style_header(ws_slices, 1, len(slice_headers))

        num_slices = predictions.shape[2]
        for si in range(num_slices):
            slice_mask = predictions[:, :, si]
            slice_stats = get_prediction_statistics(slice_mask, voxel_dims, scale_factors)
            for s in slice_stats:
                if s["class_id"] == 0:
                    continue
                row = [si, s["class_name"], s["pixel_count"], s["percentage"]]
                if has_volume:
                    row.extend([s.get("volume_mm3", ""), s.get("volume_cm3", "")])
                ws_slices.append(row)

        for row in ws_slices.iter_rows(min_row=2, max_col=len(slice_headers)):
            for cell in row:
                cell.border = thin_border
        auto_width(ws_slices)

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
